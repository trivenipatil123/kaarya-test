import openpyxl
from KaAryaProperties.settings import BASE_DIR
from core_app.models import Property


def import_property_data():
    file_path = str(BASE_DIR) + "/excel_data_files/property_data.xlsx"
    dataframe = openpyxl.load_workbook(file_path)
    dataframe1 = dataframe.active
    property_field = ['property_for', 'property_type', 'furniture_type', 'name', 'short_desc', 'long_desc', 'location',
                      'price', 'address', 'landmark', 'no_of_bed', 'no_of_bath', 'no_of_balcony', 'carper_area',
                      'builtup_area', 'super_builtup_area', 'no_of_floor', 'floor', 'allowed_for', 'facing',
                      'type_of_ownership', 'age_of_construction', 'builder', 'rera_id', 'lifts', 'parking', 'property_contact']

    for row in range(1, dataframe1.max_row):
        i, property_dict = 0, {}
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            if property_field[i] == 'parking' and not col[row].value:
                property_dict[property_field[i]] = False
            elif (property_field[i] == 'floor' or property_field[i] == 'no_of_floor' or property_field[i] == 'lifts' or
                  property_field[i] == 'no_of_balcony') and not col[row].value:
                property_dict[property_field[i]] = 0
            elif (property_field[i] == 'no_of_bed' or property_field[i] == 'no_of_bath') and not col[row].value:
                property_dict[property_field[i]] = 1
            elif property_field[i] == 'property_for' and not col[row].value:
                property_dict[property_field[i]] = 'buy'
            elif property_field[i] == 'property_type' and not col[row].value:
                property_dict[property_field[i]] = 'new'
            elif property_field[i] == 'furniture_type' and not col[row].value:
                property_dict[property_field[i]] = 'unfurnished'
            else:
                property_dict[property_field[i]] = col[row].value
            i += 1
        print(property_dict, )
        property_record = Property(**property_dict)
        property_record.save()
        print(property_record)

